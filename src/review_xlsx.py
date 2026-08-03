"""Render the weekly review's per-staff worklists (data/review.build_staff_workbook)
to an .xlsx workbook — one TAB PER STAFF MEMBER.

openpyxl is already a dependency (pandas reads the Karbon Excel exports through it),
so this adds no package to the locked-down server — the same reasoning that put
review_pdf.py on matplotlib. We drive openpyxl directly rather than going through
pandas.to_excel: each sheet is not one flat table but a stack of work-type blocks,
each with its own banded header, which a DataFrame dump can't express.

Each sheet is that person's open work — client, title, status, work type, start,
due, days overdue, days open — grouped into work-type blocks, worst overdue at the
top of each block. Work type is a literal copy of the export's own Work Type cell;
there is no project-level type column (see review._xlsx_row for why). Dates are
written as real dates (not text) so Excel sorts and filters them properly. Output
is raw .xlsx bytes, ready to stream as a download.

Readability choices (these are the point of the styling, not decoration):
  * Georgia is kept ONLY for the sheet title and the work-type banners — the brand
    voice — while every DATA cell is Calibri. A serif at 10pt over hundreds of dense
    rows is what makes a spreadsheet tiring to read; Calibri was designed for exactly
    this, and it ships with Office/Windows so no font has to be installed.
  * Dates read "Apr 15, 2026", not "2026-04-15" — still real dates underneath, so
    Excel sorts and filters them as dates.
  * Days overdue is graduated instead of uniformly red: on real data most rows are
    overdue, so painting them all red makes the column one flat wall of alarm.
    90+ days is red, 30-89 amber, 1-29 plain ink, 0 quiet grey — the eye finds the
    genuinely bad rows immediately.
  * Excel's own grey gridlines are switched OFF and replaced by one soft hairline
    under each row, plus faint zebra banding, so the horizontal scan across nine
    columns doesn't lose its place.

Brand: the firm's letterhead navy + green, echoing review_pdf.py so a printed sheet
looks of a piece with the PDF.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# Letterhead palette, softened for long reading. openpyxl wants RGB hex with no '#'.
NAVY = "1A3F8F"       # banner fill + title ink (brand)
GREEN_INK = "3D7A42"  # tab colour for a staff member with nothing overdue
INK = "1F2937"        # body text: dark slate, not pure black (less glare)
MUTED = "6B7280"      # subtitle, and the "0 days overdue" figures
BAND = "E8EDF5"       # cool pale navy for the column-header band
ZEBRA = "F7F9FC"      # barely-there banding on alternate data rows
RULE = "DCE3EC"       # soft cool hairline under each row (was a warm tan)
GRID = "C7D3E2"       # column separators: a touch stronger than the row hairline,
                      # because 8 verticals next to hundreds of horizontals read as
                      # weaker than they are if both are given the same weight
SEV_HIGH = "B91C1C"   # 90+ days overdue
SEV_MID = "B45309"    # 30-89 days overdue
WHITE = "FFFFFF"

SERIF = "Georgia"     # title + block banners only
SANS = "Calibri"      # every data cell
DATE_FMT = "mmm d, yyyy"

# Days-overdue thresholds for the graduated colour above.
SEV_HIGH_DAYS = 90
SEV_MID_DAYS = 30

# Excel's own limits: a sheet title may not exceed 31 characters and may not
# contain any of these, nor start/end with an apostrophe.
SHEET_NAME_MAX = 31
ILLEGAL_SHEET_CHARS = set("[]:*?/") | {"\\"}

# (header, row key, width, right-aligned?) — one entry per column, in sheet order.
COLUMNS = [
    ("Client",            "client",       34, False),
    ("Work / project",    "title",        42, False),
    # Karbon's longest real statuses run ~38 chars ("Ready To Start - Resend client
    # requests"); a narrower column clips them, since the neighbouring cell is
    # always populated so Excel can't overflow into it.
    ("Status",            "status",       40, False),
    ("Work type",         "work_type",    24, False),
    ("Start date",        "start_date",   14, True),
    ("Due date",          "due_date",     14, True),
    ("Days overdue",      "days_overdue", 13, True),
    ("Days open",         "days_open",    11, True),
    ("Assignee",          "assignee",     24, False),
]

# Styles are built ONCE and shared across cells: a fresh Font/Fill per cell would
# bloat the workbook's style table and slow a 1000-row export for no benefit.
_S_ROW = Side(style="thin", color=RULE)
_S_COL = Side(style="thin", color=GRID)
# One border per column position: every cell closes on its RIGHT, and the first
# column also draws its LEFT, so the block reads as a finished grid with no doubled
# lines. Excel's own gridlines are off, so these are the only lines on the sheet.
_BORDERS = [
    Border(bottom=_S_ROW, right=_S_COL, left=_S_COL) if i == 0
    else Border(bottom=_S_ROW, right=_S_COL)
    for i in range(len(COLUMNS))
]
_F_TITLE = Font(name=SERIF, size=14, bold=True, color=NAVY)
_F_SUBTITLE = Font(name=SANS, size=9.5, color=MUTED)
_F_BANNER = Font(name=SERIF, size=11.5, bold=True, color=WHITE)
_F_HEADER = Font(name=SANS, size=9.5, bold=True, color=INK)
_F_BODY = Font(name=SANS, size=12, color=INK)
_F_CLIENT = Font(name=SANS, size=12, bold=True, color=INK)
_F_NOTE = Font(name=SANS, size=12, italic=True, color=MUTED)
_F_OVERDUE = {
    "high": Font(name=SANS, size=12, bold=True, color=SEV_HIGH),
    "mid": Font(name=SANS, size=12, bold=True, color=SEV_MID),
    "low": Font(name=SANS, size=12, bold=True, color=INK),
    "none": Font(name=SANS, size=12, color=MUTED),
}
_FILL_BANNER = PatternFill("solid", fgColor=NAVY)
_FILL_HEADER = PatternFill("solid", fgColor=BAND)
_FILL_ZEBRA = PatternFill("solid", fgColor=ZEBRA)
_AL_LEFT = Alignment(horizontal="left", vertical="center")
_AL_RIGHT = Alignment(horizontal="right", vertical="center")
_AL_HEADER_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_AL_HEADER_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_AL_BANNER = Alignment(horizontal="left", vertical="center", indent=1)

ROW_H = 19.0         # a touch of air between rows; bumped up alongside the 12pt data font
BANNER_H = 22.0
HEADER_H = 18.0


def _severity(days_overdue: int) -> str:
    if days_overdue >= SEV_HIGH_DAYS:
        return "high"
    if days_overdue >= SEV_MID_DAYS:
        return "mid"
    return "low" if days_overdue > 0 else "none"


def _sheet_name(raw: str, used: set) -> str:
    """A staff name as a legal, unique Excel sheet title.

    Excel rejects []:*?/\\ and titles over 31 chars, and openpyxl raises outright on
    a duplicate title — two staff members whose names only differ past character 31
    would otherwise crash the whole export, so collisions get a ~2 / ~3 suffix.
    """
    name = "".join("-" if ch in ILLEGAL_SHEET_CHARS else ch for ch in (raw or "").strip())
    name = name.strip("'").strip() or "Unnamed"
    name = name[:SHEET_NAME_MAX]
    if name.casefold() not in used:      # Excel treats sheet names case-insensitively
        used.add(name.casefold())
        return name
    for n in range(2, 1000):
        suffix = f"~{n}"
        candidate = name[:SHEET_NAME_MAX - len(suffix)] + suffix
        if candidate.casefold() not in used:
            used.add(candidate.casefold())
            return candidate
    raise ValueError(f"cannot make a unique sheet name for {raw!r}")


def _gen_label(generated_at) -> str:
    """The generated-at stamp, in the same friendly style as the date cells
    ("Jul 30, 2026 at 2:05 PM"). Day/hour are un-padded by hand rather than with
    strftime's %#d / %-d, which are platform-specific and raise on the wrong OS."""
    if not isinstance(generated_at, date):      # datetime is a subclass of date
        return str(generated_at or "")
    stamp = f"{generated_at:%b} {generated_at.day}, {generated_at.year}"
    if isinstance(generated_at, datetime):
        hour = generated_at.hour % 12 or 12
        stamp += f" at {hour}:{generated_at:%M} {generated_at:%p}"
    return stamp


def _write_title(ws, sheet: dict, firm_name: str, generated_at) -> None:
    """The two heading lines above the first work-type block."""
    ws.cell(row=1, column=1, value=f"Weekly Review — {sheet['assignee']}").font = _F_TITLE
    subtitle = (f"{firm_name} · generated {_gen_label(generated_at)} · "
                f"{sheet['open_count']} open · {sheet['overdue_count']} overdue · "
                f"completed work excluded")
    ws.cell(row=2, column=1, value=subtitle).font = _F_SUBTITLE
    ws.row_dimensions[1].height = 21


def _write_block(ws, row: int, group: dict) -> int:
    """One work-type block: navy banner, column headers, then the rows (already
    sorted worst-overdue first). Returns the next free row."""
    ncols = len(COLUMNS)
    overdue = sum(1 for r in group["rows"] if r["days_overdue"] > 0)
    banner = (f"{group['work_type']}   —   {len(group['rows'])} open"
              + (f", {overdue} overdue" if overdue else ", none overdue"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=banner)
    cell.font = _F_BANNER
    cell.fill = _FILL_BANNER
    cell.alignment = _AL_BANNER
    ws.row_dimensions[row].height = BANNER_H
    row += 1

    for col, (header, _key, _w, right) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _F_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _AL_HEADER_RIGHT if right else _AL_HEADER_LEFT
        cell.border = _BORDERS[col - 1]
    ws.row_dimensions[row].height = HEADER_H
    row += 1

    for i, r in enumerate(group["rows"]):
        zebra = i % 2 == 1
        severity = _severity(r["days_overdue"])
        for col, (_header, key, _w, right) in enumerate(COLUMNS, start=1):
            value = r.get(key)
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = _AL_RIGHT if right else _AL_LEFT
            cell.border = _BORDERS[col - 1]
            if zebra:
                cell.fill = _FILL_ZEBRA
            if isinstance(value, date):
                cell.number_format = DATE_FMT
            if key == "days_overdue":
                cell.font = _F_OVERDUE[severity]
            elif key == "client":
                cell.font = _F_CLIENT
            else:
                cell.font = _F_BODY
        ws.row_dimensions[row].height = ROW_H
        row += 1

    return row + 1      # one blank spacer row between blocks


def _write_sheet(ws, sheet: dict, firm_name: str, generated_at) -> None:
    _write_title(ws, sheet, firm_name, generated_at)
    row = 4
    for group in sheet["groups"]:
        row = _write_block(ws, row, group)
    if not sheet["groups"]:
        ws.cell(row=row, column=1, value="No open work in scope.").font = _F_NOTE

    for col, (_header, _key, width, _right) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    # Freeze the heading lines ONLY (rows 1-3). Deliberately NOT "B4": freezing a
    # column makes Excel draw a heavy black split line down the whole sheet, over the
    # title and straight through the navy banners — the user flagged it. Column
    # separation comes from the cell borders instead.
    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = GREEN_INK if not sheet["overdue_count"] else NAVY
    # Excel's default grey grid over nine wide columns is visual noise; the row
    # hairlines and zebra banding already carry the eye across.
    ws.sheet_view.showGridLines = False
    # Print sanely by default: landscape, all nine columns squeezed onto one page
    # wide. sheet_properties.pageSetUpPr is None until we give it one.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def render_xlsx(wb_data: dict) -> bytes:
    """The workbook from build_staff_workbook as .xlsx bytes — one sheet per staff
    member, in the order the builder put them (most overdue first, Unassigned last)."""
    wb = Workbook()
    wb.remove(wb.active)        # drop the default "Sheet"; we name every tab ourselves

    firm_name = wb_data.get("firm_name", "")
    generated_at = wb_data.get("generated_at")
    used: set = set()
    for sheet in wb_data.get("sheets", []):
        _write_sheet(wb.create_sheet(_sheet_name(sheet["assignee"], used)),
                     sheet, firm_name, generated_at)

    if not wb.sheetnames:       # a workbook must have at least one sheet
        ws = wb.create_sheet("Weekly Review")
        ws.cell(row=1, column=1, value="Weekly Review").font = _F_TITLE
        ws.cell(row=2, column=1,
                value=f"{firm_name} · generated {_gen_label(generated_at)}").font = _F_SUBTITLE
        ws.cell(row=4, column=1,
                value="No open work in scope — nothing to export.").font = _F_NOTE
        ws.column_dimensions["A"].width = 60
        ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
