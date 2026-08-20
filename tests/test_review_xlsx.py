"""Tests for review_xlsx.render_xlsx (the .xlsx writer) and the /review.xlsx route.

The writer is checked by round-tripping the bytes back through openpyxl, so the
assertions are about what Excel will actually see: sheet names, cell values, and
real dates rather than date-looking text.
"""
import io
from datetime import date, datetime

import pytest
from openpyxl import load_workbook

import review_xlsx
import webapp
from data.store import Store

GEN = datetime(2026, 7, 30, 9, 15)


def _row(client, title, days_overdue, work_type="Tax: 1040", status="In Progress",
         assignee="Sarah", due=date(2026, 4, 15), start=date(2026, 2, 1)):
    return {"client": client, "title": title, "status": status,
            "work_type": work_type, "start_date": start, "due_date": due,
            "days_overdue": days_overdue, "days_open": 120, "assignee": assignee}


def _book(sheets):
    return {"firm_name": "NLC Financial", "generated_at": GEN, "sheets": sheets}


def _sheet(assignee, groups):
    rows = [r for g in groups for r in g["rows"]]
    return {"assignee": assignee, "open_count": len(rows),
            "overdue_count": sum(1 for r in rows if r["days_overdue"] > 0),
            "groups": groups}


def _group(work_type, rows):
    return {"work_type": work_type,
            "max_days_overdue": max((r["days_overdue"] for r in rows), default=0),
            "rows": rows}


def _load(book):
    data = review_xlsx.render_xlsx(book)
    assert data[:2] == b"PK", "not a zip container — Excel would reject it"
    return load_workbook(io.BytesIO(data))


# ---- The writer ----------------------------------------------------------

def test_renders_one_sheet_per_staff_member_in_builder_order():
    wb = _load(_book([
        _sheet("Sarah", [_group("Tax: 1040", [_row("Acme", "1040 Return", 106)])]),
        _sheet("James", [_group("Payroll", [_row("Beta", "Q1 Payroll", 12)])]),
        _sheet("Unassigned", [_group("Bookkeeping", [_row("Ceta", "Jan books", 3)])]),
    ]))
    assert wb.sheetnames == ["Sarah", "James", "Unassigned"]


def test_block_banner_then_headers_then_rows():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106),
                             _row("Beta", "State Return", 92)]),
    ])]))
    ws = wb["Sarah"]
    assert ws["A1"].value == "Weekly Review — Sarah"
    assert "2 open" in ws["A2"].value and "completed work excluded" in ws["A2"].value
    assert ws["A4"].value.startswith("Tax: 1040")        # the merged navy banner
    assert [c.value for c in ws[5]][:4] == ["Client", "Work / project", "Status",
                                            "Work type"]
    assert [ws.cell(row=r, column=1).value for r in (6, 7)] == ["Acme", "Beta"]


def test_dates_are_written_as_real_dates_not_text():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    ws = wb["Sarah"]
    start, due = ws.cell(row=6, column=5), ws.cell(row=6, column=6)
    assert start.value == datetime(2026, 2, 1)      # openpyxl reads dates as datetime
    assert due.value == datetime(2026, 4, 15)
    assert due.number_format == review_xlsx.DATE_FMT


def test_blank_due_date_leaves_the_cell_empty():
    wb = _load(_book([_sheet("Sarah", [
        _group("Unclassified", [_row("Acme", "W-2", 26, due=None)]),
    ])]))
    assert wb["Sarah"].cell(row=6, column=6).value is None


def test_every_column_is_written_in_order():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    ws = wb["Sarah"]
    assert [c.value for c in ws[5]] == [h for h, _k, _w, _r in review_xlsx.COLUMNS]
    # No "Project type": the export carries the document's OWN Work Type only.
    assert [c.value for c in ws[5]] == [
        "Client", "Work / project", "Status", "Work type", "Start date", "Due date",
        "Days overdue", "Days open", "Assignee"]
    assert [ws.cell(row=6, column=c).value for c in (1, 2, 3, 4, 7, 8, 9)] == [
        "Acme", "1040 Return", "In Progress", "Tax: 1040", 106, 120, "Sarah"]


def test_blocks_are_separated_and_each_carries_its_own_headers():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
        _group("Payroll", [_row("Beta", "Q1 Payroll", 12, work_type="Payroll")]),
    ])]))
    ws = wb["Sarah"]
    assert ws["A4"].value.startswith("Tax: 1040")
    assert ws["A5"].value == "Client"
    assert ws["A6"].value == "Acme"
    assert ws["A7"].value is None                 # spacer row between blocks
    assert ws["A8"].value.startswith("Payroll")
    assert ws["A9"].value == "Client"
    assert ws["A10"].value == "Beta"


def test_only_the_heading_rows_are_frozen_never_a_column():
    # Freezing a COLUMN makes Excel draw a heavy black split line down the entire
    # sheet, over the title and through the navy banners. Rows only.
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    assert wb["Sarah"].freeze_panes == "A4"


def test_every_column_is_separated_within_each_block():
    # Column separators are drawn PER CELL, so each work-type block carries its own
    # grid and the lines stop at the banner and the spacer row -- rather than one
    # continuous rule running the height of the sheet.
    wb = _load(_book([
        _sheet("Sarah", [
            _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
            _group("Payroll", [_row("Beta", "Q1 Payroll", 12, work_type="Payroll")]),
        ]),
    ]))
    ws = wb["Sarah"]
    for row in (5, 6, 9, 10):        # header + data of BOTH blocks
        for col in range(1, len(review_xlsx.COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.right.style == "thin", f"{cell.coordinate} right edge"
            assert cell.border.bottom.style == "thin", f"{cell.coordinate} bottom edge"
        assert ws.cell(row=row, column=1).border.left.style == "thin"
    # The banner and the spacer between blocks carry no verticals, so each block's
    # grid is visually self-contained.
    assert ws.cell(row=4, column=3).border.right.style is None      # banner
    assert ws.cell(row=7, column=3).border.right.style is None      # spacer


def test_status_column_is_wide_enough_for_karbons_longest_statuses():
    # "Ready To Start - Resend client requests" clipped at the old width, and the
    # neighbouring cell is always populated so Excel cannot overflow into it.
    width = next(w for h, _k, w, _r in review_xlsx.COLUMNS if h == "Status")
    assert width >= len("Ready To Start - Resend client requests")


def test_banner_counts_overdue_within_the_block():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106),
                             _row("Beta", "State Return", 0)]),
    ])]))
    assert wb["Sarah"]["A4"].value == "Tax: 1040   —   2 open, 1 overdue"


def test_block_with_nothing_overdue_says_so():
    wb = _load(_book([_sheet("Sarah", [
        _group("Payroll", [_row("Acme", "Q1 Payroll", 0, work_type="Payroll")]),
    ])]))
    assert wb["Sarah"]["A4"].value.endswith("1 open, none overdue")


def test_empty_workbook_still_saves_with_a_placeholder_sheet():
    # openpyxl cannot save a workbook with zero sheets -- a quiet week must not
    # turn the export into a 500.
    wb = _load(_book([]))
    assert wb.sheetnames == ["Weekly Review"]
    assert "No open work in scope" in wb["Weekly Review"]["A4"].value


def test_staff_member_with_no_groups_gets_a_note_not_an_empty_sheet():
    wb = _load(_book([_sheet("Sarah", [])]))
    assert wb["Sarah"]["A4"].value == "No open work in scope."


# ---- Readability styling -------------------------------------------------

def test_dates_show_month_day_year_not_iso():
    # "Apr 15, 2026", not "2026-04-15" -- still a real date underneath.
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    assert review_xlsx.DATE_FMT == "mmm d, yyyy"
    assert wb["Sarah"].cell(row=6, column=6).number_format == "mmm d, yyyy"


def test_generated_stamp_reads_as_a_friendly_date():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    # GEN is 2026-07-30 09:15 -- no zero-padded day, no 24h clock, no ISO.
    assert "Jul 30, 2026 at 9:15 AM" in wb["Sarah"]["A2"].value


def test_data_cells_are_sans_and_headings_stay_serif():
    # Georgia is the brand voice for the title/banner; the dense data grid is
    # Calibri, which is what makes hundreds of rows readable.
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    ws = wb["Sarah"]
    assert ws["A1"].font.name == review_xlsx.SERIF == "Georgia"       # title
    assert ws["A4"].font.name == review_xlsx.SERIF                    # block banner
    assert ws["A5"].font.name == review_xlsx.SANS == "Calibri"        # column header
    assert ws["A6"].font.name == review_xlsx.SANS                     # data
    assert ws.cell(row=6, column=7).font.name == review_xlsx.SANS     # days overdue


def test_days_overdue_colour_is_graduated_not_uniformly_red():
    # Most rows are overdue on real data, so painting them all red makes the
    # column one flat wall of alarm and hides the genuinely bad ones.
    wb = _load(_book([_sheet("Sarah", [_group("Tax: 1040", [
        _row("Acme", "a", 200),     # 90+   -> red
        _row("Beta", "b", 45),      # 30-89 -> amber
        _row("Ceta", "c", 5),       # 1-29  -> plain ink
        _row("Delta", "d", 0),      # 0     -> quiet grey
    ])])]))
    ws = wb["Sarah"]
    colours = [ws.cell(row=r, column=7).font.color.rgb for r in (6, 7, 8, 9)]
    assert [c[-6:] for c in colours] == [
        review_xlsx.SEV_HIGH, review_xlsx.SEV_MID, review_xlsx.INK, review_xlsx.MUTED]
    # The 0-overdue figure is also the only one not bolded.
    assert [ws.cell(row=r, column=7).font.bold for r in (6, 7, 8, 9)][-1] is not True


def test_severity_thresholds():
    assert review_xlsx._severity(90) == "high"
    assert review_xlsx._severity(89) == "mid"
    assert review_xlsx._severity(30) == "mid"
    assert review_xlsx._severity(29) == "low"
    assert review_xlsx._severity(1) == "low"
    assert review_xlsx._severity(0) == "none"


def test_excels_own_gridlines_are_turned_off():
    wb = _load(_book([_sheet("Sarah", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 106)]),
    ])]))
    assert wb["Sarah"].sheet_view.showGridLines is False


def test_alternate_data_rows_are_banded():
    wb = _load(_book([_sheet("Sarah", [_group("Tax: 1040", [
        _row("Acme", "a", 200), _row("Beta", "b", 100), _row("Ceta", "c", 50),
    ])])]))
    ws = wb["Sarah"]
    fills = [ws.cell(row=r, column=1).fill.fgColor.rgb for r in (6, 7, 8)]
    assert fills[1][-6:] == review_xlsx.ZEBRA        # second row banded
    assert fills[0][-6:] != review_xlsx.ZEBRA        # first and third are not
    assert fills[2][-6:] != review_xlsx.ZEBRA


# ---- Sheet-name sanitizing ----------------------------------------------

def test_illegal_sheet_characters_are_replaced():
    # Excel rejects []:*?/\ in a sheet title, and a client-style name like
    # "Smith/Jones" would otherwise blow up the whole export.
    wb = _load(_book([_sheet("Smith/Jones: A*B?[C]", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 5)]),
    ])]))
    assert wb.sheetnames == ["Smith-Jones- A-B--C-"]


def test_long_names_are_truncated_to_excels_31_char_limit():
    wb = _load(_book([_sheet("Bartholomew Fitzwilliam Montgomery Esquire", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 5)]),
    ])]))
    assert wb.sheetnames == ["Bartholomew Fitzwilliam Montgom"]
    assert len(wb.sheetnames[0]) == review_xlsx.SHEET_NAME_MAX == 31


def test_names_colliding_after_truncation_are_deduped():
    # Two real staff members whose names only differ past character 31 would
    # otherwise make openpyxl raise on the duplicate title.
    long_a = "Bartholomew Fitzwilliam Montgomery Senior"
    long_b = "Bartholomew Fitzwilliam Montgomery Junior"
    wb = _load(_book([
        _sheet(long_a, [_group("Tax: 1040", [_row("Acme", "1040 Return", 5)])]),
        _sheet(long_b, [_group("Tax: 1040", [_row("Beta", "1040 Return", 4)])]),
    ]))
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "Bartholomew Fitzwilliam Montgom"
    assert wb.sheetnames[1] == "Bartholomew Fitzwilliam Montg~2"
    assert wb["Bartholomew Fitzwilliam Montg~2"]["A6"].value == "Beta"


def test_blank_assignee_name_still_gets_a_usable_sheet():
    wb = _load(_book([_sheet("   ", [
        _group("Tax: 1040", [_row("Acme", "1040 Return", 5)]),
    ])]))
    assert wb.sheetnames == ["Unnamed"]


# ---- The route -----------------------------------------------------------

@pytest.fixture
def client(tmp_path, monkeypatch):
    store = Store(tmp_path / "test.db")
    monkeypatch.setattr(webapp, "_store", store)
    webapp.app.testing = True
    with webapp.app.test_client() as c:
        with c.session_transaction() as sess:
            sess["csrf_token"] = "test-token"
        yield c
    store.close()


def test_route_serves_a_downloadable_workbook(client):
    r = client.get("/review.xlsx")
    assert r.status_code == 200
    assert r.mimetype == webapp.XLSX_MIMETYPE
    disposition = r.headers["Content-Disposition"]
    assert disposition.startswith("attachment;")
    assert disposition.endswith(f'filename="weekly-review-{date.today().isoformat()}.xlsx"')
    # Real bytes Excel can open. Deliberately not asserting the exact sheet list:
    # _get_data() memoizes on (data_version, today, overdue_days), so a fresh store
    # can legitimately hit another test's cache entry.
    assert load_workbook(io.BytesIO(r.data)).sheetnames


def test_route_needs_no_session(client):
    """The dashboard has no sign-in: a visitor with an empty session still gets
    the workbook. This replaces an older test that asserted the opposite, back
    when the route sat behind an admin login."""
    # Same monkeypatched temp store as the fixture -- an un-patched app would build
    # (and write to) the REAL app-data database just to answer this request.
    with client.session_transaction() as sess:
        sess.clear()
    r = client.get("/review.xlsx")
    assert r.status_code == 200
    assert r.mimetype == webapp.XLSX_MIMETYPE
